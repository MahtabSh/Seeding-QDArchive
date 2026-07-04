#!/usr/bin/env python3
"""
Generate XLSX classification export for Part 2 Step 4c of SQ26.

Columns: repository_id, project_type, project_title,
         primary_class, secondary_class, no_project_files

Output: 23692652-sq26-classification.xlsx
"""
import sqlite3
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Install: pip install openpyxl")
    raise

DB = "23692652-sq26.db"
OUT = "23692652-sq26-classification.xlsx"

HEADER_COLOR = "1F4E79"   # dark blue
ALT_COLOR    = "DEEAF1"   # light blue alternating rows


def main() -> None:
    conn = sqlite3.connect(DB)

    rows = conn.execute("""
        SELECT
            p.repository_id,
            COALESCE(p.type, 'UNKNOWN') AS project_type,
            p.title AS project_title,
            CASE
                WHEN cv.section IS NOT NULL
                THEN cv.section || ' - ' || COALESCE(cv.section_name, '')
                ELSE ''
            END AS primary_class,
            CASE
                WHEN cv.division IS NOT NULL AND cv.division != ''
                THEN cv.division || ' - ' || COALESCE(cv.division_name, '')
                ELSE ''
            END AS secondary_class,
            COALESCE(fc.file_count, 0) AS no_project_files
        FROM projects p
        LEFT JOIN classifications_vote cv ON p.id = cv.project_id
        LEFT JOIN (
            SELECT project_id, COUNT(*) AS file_count
            FROM files
            GROUP BY project_id
        ) fc ON p.id = fc.project_id
        ORDER BY p.repository_id, p.type, p.id
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classifications"

    headers = [
        "repository_id",
        "project_type",
        "project_title",
        "primary_class",
        "secondary_class",
        "no_project_files",
    ]

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR,
                               fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    alt_fill     = PatternFill(start_color=ALT_COLOR, end_color=ALT_COLOR,
                               fill_type="solid")
    thin_side    = Side(style="thin", color="CCCCCC")
    border       = Border(bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20

    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            cell.border = border
            if col_idx == 3:   # title column: left-align
                cell.alignment = Alignment(horizontal="left")
            elif col_idx in (4, 5):
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    col_widths = [16, 18, 55, 55, 55, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows)+1}"

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Repository", "Project Type", "Count"])
    for row in conn.execute(
        "SELECT repository_id, type, COUNT(*) FROM projects "
        "GROUP BY repository_id, type ORDER BY repository_id, type"
    ) if False else []:
        pass  # handled below

    conn2 = sqlite3.connect(DB)
    sum_rows = conn2.execute(
        "SELECT repository_id, COALESCE(type,'UNKNOWN'), COUNT(*) "
        "FROM projects GROUP BY repository_id, type ORDER BY repository_id, type"
    ).fetchall()
    conn2.close()

    ws_sum.cell(1, 1, "repository_id").font = Font(bold=True)
    ws_sum.cell(1, 2, "project_type").font  = Font(bold=True)
    ws_sum.cell(1, 3, "count").font         = Font(bold=True)
    for r in sum_rows:
        ws_sum.append(list(r))

    wb.save(OUT)
    print(f"Saved {len(rows):,} rows → {OUT}")


if __name__ == "__main__":
    main()
